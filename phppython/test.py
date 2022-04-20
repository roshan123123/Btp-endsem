
#wb= Workbook()  to create a new workbook
 # will give the active worksheet
#ws.title="data" changing the name of title
# print(ws)   
# print(ws['A1'].value) accesing the value
# ws['A2'].value="sjijcn" and then save it to get the chnges done by 
# wb.save('name where we want to save it').
# ws['A2']="sjijcn"  this also works fine
#print(wb.sheetnames)  prints the sheet name
#wb['nameofset'] -while choosing ws other than active
#wb.create_sheet("name with extension")  to create sheet

#print(get_column_letter(27))
#print(ws[char+str(rownumber)])

#VVVVVVVI
#accesing the data ws.cell(3,4).value /// row,col in cell
#row=sh1.max_row
#colmn=sh1.max_column

import sys
import math
res=sys.argv[1]
inputfile=res
Rolls_sorted="Rolls_sorted_"+ inputfile
Grades_sorted="Grades_sorted_"+ inputfile
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook  #to load an existing workbook
from openpyxl.utils import get_column_letter
wb =load_workbook(inputfile)  #abs path alwyas works and loaded workbook
ws =wb.active 
student_count=ws.max_row-3
assesment_count=ws.max_column-3
def merge():
    ws.merge_cells(f"{get_column_letter(4+assesment_count)}1:{get_column_letter(3+2*assesment_count)}1")
    ws.cell(1,4+assesment_count).value="Out of 100 for each assesment"
   # ws.merge_cells()

def copy_assesment_name():
    for col in range (1,assesment_count+1):
        ws.cell(2,3+assesment_count+col).value=ws.cell(2,3+col).value+" out of 100"

def copy_100():
     for col in range (1,assesment_count+2):
        ws.cell(3,3+assesment_count+col).value=100

def copy_total_diff_grades():
    ws.cell(1,4+2*assesment_count).value="Total"
    ws.cell(2,4+2*assesment_count).value=100
    ws.cell(2,5+2*assesment_count).value="Difference"
    ws.cell(2,6+2*assesment_count).value="Grades"

def  calculate_out_of_100():
    for col in range(1,assesment_count+1):
        for row in range(1,student_count+1):
            actualrow=row+3
            actualcol=col+3+assesment_count
            refrow=3+row
            refcol=col+3
            totalmarks=ws.cell(3,refcol).value
            
            #here we can make it formula based also scope is there 
            ws.cell(actualrow,actualcol).value=round((ws.cell(refrow,refcol).value *100)/totalmarks,1)

def calculate_total():
    #here also scope to feed formula int xlsx file
    targetcol=3+assesment_count*2+1
    for row in range (1,student_count+1):
        tsum=0
        for col in range(1,assesment_count+1):
            tsum+=ws.cell(row+3,3+assesment_count+col).value*ws.cell(1,3+col).value/100

        ws.cell(row+3,targetcol).value=round(tsum,1)
def calculate_diff(): 
    wb1=load_workbook(Grades_sorted)
    ws1=wb1.active
    for row in range(2,student_count+1):
        ws1.cell(row+3,3+2*assesment_count+2).value=ws1.cell(row+2,3+2*assesment_count+1).value-ws1.cell(row+3,3+2*assesment_count+1).value

    wb1.save(Grades_sorted)   


def calculate_grades():
    wb1=load_workbook(Grades_sorted)
    ws1=wb1.active
    #making the list for grades
    percent=[5,15,25,30,15,5,5]
    dic={0:"AA",1:"AB",2:"BB",3:"BC",4:"CC",5:"CD",6:"DD"}
    grade=[]
    for i in range(0,7):
        x=math.ceil(student_count*percent[i]/100)
        while x>0:
            grade.append(dic[i])
            x=x-1
    p=student_count
   
    for row in range (1,student_count+1):
        ws1.cell(3+row,3+2*assesment_count+3).value=grade[row-1]

    wb1.save(Grades_sorted)  


def  make_table_of_grade_count(filename):
    wb1=load_workbook(filename)
    ws1=wb1.active
    ws1.cell(3,6+2*assesment_count+1).value="Stud Total" 
    ws1.cell(3,6+2*assesment_count+2).value=student_count
    ws1.cell(4,6+2*assesment_count+1).value="Grade" 
    ws1.cell(4,6+2*assesment_count+2).value="IAPC Recom per"
    ws1.cell(4,6+2*assesment_count+3).value="IAPC Recom count" 
    ws1.cell(4,6+2*assesment_count+4).value="This Sheet count"
    dic={"AA":5,"AB":15,"BB":25,"BC":30,"CC":15,"CD":5,"DD":5}
    row=1
    for i in dic:
        ws1.cell(4+row,6+2*assesment_count+1).value=i
        ws1.cell(4+row,6+2*assesment_count+2).value=dic[i]
        row+=1
    #now populating coun accrdng to ipac
    for row in range(1,8):
        ws1.cell(4+row,6+2*assesment_count+3).value=round(student_count*ws1.cell(4+row,6+2*assesment_count+2).value/100)
        st=str(get_column_letter(6+2*assesment_count))+"4"
        en=str(get_column_letter(6+2*assesment_count))+str(3+student_count)
        valcell=str(get_column_letter(6+2*assesment_count+1))+str(4+row)
        ws1.cell(4+row,6+2*assesment_count+4).value=f"=COUNTIF({st}:{en},{valcell})"
    

    #populate the count of grades in our sheet 

    wb1.save(filename) 


def make_Roll_sorted(col):
    ##make Roll sorted
    wb1=load_workbook(Grades_sorted)
    ws1=wb1.active
    #res = [((i), (i + 1) % len(test_list)) 
        # for row in range(len(test_list))]
    #shift up 
    #populate with grade count thing
    #save it with the name of Roll_sorted


def correct_the_index(name):
    wb1=load_workbook(name)
    ws1=wb1.active
    for row in range(1,student_count+1):
        ws1.cell(3+row,1).value=row
    wb1.save(name)    

    
def solve():
   
    merge() #merge the column saying each assesment out of 100
    copy_assesment_name()  #name of assignment out of 100
    copy_100()    #copies 100 to that many guys
    copy_total_diff_grades() #naming of coulumn
    calculate_out_of_100()    #calculate marks out of 100 for every assesment
    calculate_total()        #calculates overall out of 100
    print(wb.worksheets)
    wb.save(Grades_sorted)
    
    df=pd.read_excel(Grades_sorted,engine='openpyxl')
    
    df=df.sort_values('Total',ascending=False)
    df.to_excel(Grades_sorted,index=False)
    correct_the_index(Grades_sorted)
    

   
 
   
 
    #sort the complete sheet on the basis of total score
    calculate_diff()         #prepares difference coulumn
    calculate_grades()       #calculate grades
    df=pd.read_excel(Grades_sorted,engine='openpyxl')
    
    df=df.sort_values('Roll',ascending=True)
    df.to_excel(Rolls_sorted,index=False)
    correct_the_index(Rolls_sorted)
    #correction can be done
    #correct_the_index()

    #make_Roll_sorted(col)    #col is the column on which you want it to sort and will make roll sorted spreadsheet
    make_table_of_grade_count(Grades_sorted) #craete complete table of grade count according to the iapc rule
    make_table_of_grade_count(Rolls_sorted)
   



weightage_sum=0
for col in range(4,4+assesment_count):
    weightage_sum+=ws.cell(1,col).value

if(weightage_sum==100):
   solve()
else:
    print("check your file the weightage sum is not 100")





#sort ws on the basis of grades

#sort on the basis of serial number or roll number



